import openpyxl
import openpyxl.workbook
from openpyxl.worksheet.worksheet import Worksheet
from MyFunctions import error


class Excel:

    def __init__(self, file) -> None:
        self.records = []
        self.__workbook = None
        self.__sheet = None

        self.__LoadWorkbookAndSheet(file)

        self.__CheckIfAllAttributesAreLocatedWithinExcel(self.__sheet)

        self.__CreateListWithAllTheRecords(self.__sheet)
        pass

    def __LoadWorkbookAndSheet(self, file):
        self.__workbook = openpyxl.load_workbook(file)
        self.__sheet = self.__workbook["Test Cases"]

        assert self.__sheet != None and self.__workbook != None

    def __CreateListWithAllTheRecords(self, sheet: Worksheet):

        id = self.__Extract(sheet, "Id")
        name = self.__Extract(sheet, "Name")
        lot2id = self.__Extract(sheet, "Lot2 ID")
        automated = self.__Extract(sheet, "Automated")
        automatable = self.__Extract(sheet, "Automatable")
        created_on = self.__Extract(sheet, "Created On (Version)")
        tool = self.__Extract(sheet, "Tool")

        LISTS_HAVENT_THE_SAME_SIZE = not all(
            len(lst) == len(id)
            for lst in [name, lot2id, automated, created_on, automatable, tool]
        )

        if LISTS_HAVENT_THE_SAME_SIZE:
            error("The size of each column isnt the same")
            exit()
            pass

        for id, name, lot2id, automated, created_on, automatable, tool in zip(
            id, name, lot2id, automated, created_on, automatable, tool
        ):
            # Format of record
            record = (id, name, lot2id, automatable, automated, created_on, tool)
            self.records.append(record)
            pass

        assert len(self.records) != 0
        pass

    def __Extract(self, sheet: Worksheet, attribute):
        data = None

        for col in sheet.iter_cols(values_only=True):
            column_name = col[0].strip()

            if column_name == attribute:
                data = list(col[1:])

        data = self.__CleanTheListFromNoneValues(data)

        assert data is not None
        return data

    def __CleanTheListFromNoneValues(self, list):
        return [x for x in list if x is not None]

    def __CheckIfAllAttributesAreLocatedWithinExcel(self, sheet: Worksheet):

        attributes = [
            "Id",
            "Name",
            "Automated",
            "Lot2 ID",
            "Created On (Version)",
            "Automatable",
            "Tool",
        ]

        for col in sheet.iter_cols(values_only=True):
            column_name = col[0].strip() if col[0] is not None else None

            if column_name in attributes:
                attributes.remove(column_name)
                if len(attributes) == 0:
                    break

        if len(attributes) != 0:
            error("Attributes", attributes, "not found within excel file")
            exit()

        assert len(attributes) == 0
        pass

    def __str__(self):
        for record in self.records:
            print(record)
        return ""

    pass
