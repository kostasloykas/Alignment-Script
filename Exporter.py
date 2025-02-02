import pprint
from Collector import Collector
import openpyxl
import openpyxl.workbook
from openpyxl.worksheet.worksheet import Worksheet

from MyFunctions import debug


class Exporter:

    def __init__(self, collector: Collector) -> None:
        self.collector: Collector = None
        self.excel = None

        self.collector = collector
        self.excel = openpyxl.Workbook()
        self.excel.remove(self.excel["Sheet"])  # delete the first sheet (its unusable)

        assert self.collector is not None and self.excel is not None
        pass

    def __CreateTheSheetInWorkbook(self, name):
        self.excel.create_sheet(title=name)
        debug("Sheet", name, "created")

        assert name in self.excel.sheetnames
        pass

    def __SaveExcelFile(self):
        if len(self.excel.sheetnames) == 0:
            self.excel.create_sheet("All testacases are aligned")
        self.excel.save("Results.xlsx")
        pass

    def __AddAttributesToTheFirstRowOfTheSheet(self, sheet):
        attributes = [
            "Id",
            "Name",
            "Lot2 ID",
            "Automatable",
            "Automated",
            "Created On (Version)",
            "Tool",
            "Comments",
        ]

        # Write the column names to the first row (starting at column 1)
        for col_num, column_name in enumerate(attributes, start=1):
            self.excel[sheet].cell(row=1, column=col_num, value=column_name)
        pass

    # TODO:__AddRecordInRowOfTheSheet
    def __AddRecordInNewRowOfTheSheet(self, record, sheet):

        self.excel[sheet].append(record)
        pass

    # TODO:__AddDataToTheSheet
    def __AddDataToTheSheet(self, data, sheet):
        assert sheet in self.excel.sheetnames

        self.__AddAttributesToTheFirstRowOfTheSheet(sheet)

        if sheet == "Not-Aligned-Records":
            for record in data:
                record1, record2, comment = record
                self.__AddRecordInNewRowOfTheSheet(record1 + (comment,), sheet)
                self.__AddRecordInNewRowOfTheSheet(record2, sheet)
                self.__AddRecordInNewRowOfTheSheet([None], sheet)  # Add an empty row

        else:
            for record in data:
                self.__AddRecordInNewRowOfTheSheet(record, sheet)
                pass
        pass

    def ExportTheExcelFile(self) -> None:
        debug("Starting the export of the excel file")
        book = self.collector.book

        for sheet in book:
            self.__CreateTheSheetInWorkbook(sheet)
            self.__AddDataToTheSheet(data=book[sheet], sheet=sheet)

        self.__SaveExcelFile()
        debug("The excel file has been saved succesfully")
        pass

    pass
